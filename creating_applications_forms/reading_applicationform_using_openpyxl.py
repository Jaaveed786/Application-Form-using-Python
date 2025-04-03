from openpyxl import Workbook,load_workbook

import tkinter as tk
from tkinter import *
from tkinter import ttk
import pandas as pd
import paymentfile_application
import datetime

date=datetime.datetime.now()
dt=date.strftime('%d/%m/%y')
formatStr='%d/%m/%y'

pd.set_option('display.max_columns',500)
pd.set_option('display.max_rows',500)
pd.set_option('display.width',500)

root=tk.Tk()
root.title('  APPLICATION FORM  ')

course=['Ms Office',
        'Tally prime',
        'C Language',
        'Python',
        'Adv Python',
        'Java',
        'Adv Java',
        'DTP',
        'Photoshop',
        'Web design']

root.geometry('900x900')
root.config(bg='lightblue')
lb1=Label(root,text='APPLICATION FORM',bg='Green',font=('',35,'bold'),relief=GROOVE)
lb1.pack(fill=X)
cframe=LabelFrame(root,bg='yellow',fg='black')
cframe.pack()
lb2=Label(cframe,text='Select course : ')
cn=ttk.Combobox(cframe,width=40,values=course)

lb3=Label(cframe,text='Course fees : ')
cf=Entry(cframe,width=22)

sframe=LabelFrame(root,text='STUDENT DETAILS',fg='blue',bg='lightblue',font=('',14,'bold'))
sframe.pack()

def get_fees():
    global cr_var
    global cf_var
    cr_var=cn.get()
    cf.delete(0,END)
    if(cr_var=='Ms office'):
        cf.insert(0,1800)
        cf_var=int(cf.get())
    elif(cr_var=='Tally prime' or cr_var=='Python' or cr_var=='Java'):
        cf.insert(0,3500)
        cf_var=int(cf.get())
    elif(cr_var=='C Language'):
        cf.insert(0,3000)
        cf_var=int(cf.get())
    elif(cr_var=='Adv Python' or cr_var=='Adv Java' or cr_var=='Web design'):
        cf.insert(0,4500)
        cf_var=int(cf.get())
    elif(cr_var=='DTP'):
        cf.insert(0,4000)
        cf_var=int(cf.get())
    elif(cr_var=='Photoshop'):
        cf.insert(0,2000)
    else:
        cf.insert(0,00)
        
bu1=Button(cframe,text='Get fees',command=get_fees)



def submit():
    wb=load_workbook('data.xlsx')  ##wb=workbook
    ws=wb.active                   ##ws=worksheet 
    e1_var=e1.get()
    e2_var=e2.get()
    e3_var=e3.get()
    e4_var=e4.get()
    e5_var=e5.get()
    e6_var=e6.get()
    e7_var=e7.get('1.0','end-1c')   #to get multiple lines

    ws.append([e1_var,e2_var,e3_var,cr_var,cf_var,e4_var,e5_var,e6_var,e7_var])
    wb.save('data.xlsx')


    
bu2=Button(root,text='  Submit form  ',command=submit)

def update():
    wb=load_workbook('D:\\Python_Practice\\data.xlsx')
    ws=wb.active
    x=[]
    for row in range(1,6):
        data=[]
        for col in range(1,10):
            d=ws.cell(row,col).value
            data.append(d)
        x.append(data)

    for i in x:
        for j in range(9):
            print(i[j],end=' ')
bu3=Button(root,text= ' show ',command=update )


bu4=Button(root,text='Payment',command=paymentfile_application.pay_fees)


lb4=Label(sframe,text='Enter Student name : ')
e1=Entry(root,width=35)
e1_var=e1.get()
lb5=Label(sframe,text='Enter Father name : ')
e2=Entry(root,width=35)
e2_var=e2.get()
lb6=Label(sframe,text='Date of join : ')
e3=Entry(root,width=35)
lb7=Label(sframe,text='Qualification : ')
e4=Entry(sframe,width=35)
lb8=Label(sframe,text='Mobile number : ')
e5=Entry(sframe,width=35)
lb9=Label(sframe,text='Whatsapp number : ')
e6=Entry(sframe,width=35)
lb10=Label(sframe,text='Address : ')
e7=Text(sframe,width=35,height=6)


lb2.grid(row=0,column=0,padx=15,pady=15)
cn.grid(row=0,column=1,padx=15,pady=15)
lb3.grid(row=0,column=2,padx=15,pady=15)
cf.grid(row=0,column=3,padx=15,pady=15)
bu1.grid(row=0,column=4,padx=15,pady=15)

lb4.grid(row=1,column=0,padx=15,pady=15)
e1.grid(row=1,column=1,padx=15,pady=15)
lb5.grid(row=1,column=2,padx=15,pady=15)
e2.grid(row=1,column=3,padx=15,pady=15)
lb6.grid(row=2,column=4,padx=15,pady=15)
e3.grid(row=2,column=0,padx=15,pady=15)
lb7.grid(row=2,column=1,padx=15,pady=15)
e4.grid(row=2,column=2,padx=15,pady=15)
lb8.grid(row=3,column=3,padx=15,pady=15)
e5.grid(row=3,column=4,padx=15,pady=15)
lb9.grid(row=3,column=0,padx=15,pady=15)
e6.grid(row=3,column=1,padx=15,pady=15)
lb10.grid(row=4,column=2,padx=15,pady=15)
e7.grid(row=4,column=3,padx=15,pady=15)

bu2.place(x=400,y=570)
bu3.place(x=500,y=570)

bu4.place(x=600,y=570)
root.mainloop()
